import re, csv, os, time
import pandas as pd
import tkinter
from tkinter import filedialog
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

#rc('mathtext', default='regular')


def find_between( s, first, last ):
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""

def clear_between( s, first, last ):
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        toclear = s[start:end]
        toclear=first+toclear+last
        s=s.replace(toclear,"")
        return s
    except ValueError:
        print ("[ERROR]: Sub string didn't found")
        pass

def reformStep(s):
    s=clear_between(s,"[DEBUG] "," ")
    time=find_between(s,"",",")
    step=find_between(s,"Start command "," ")
    l=[time,step]
    return l

def reformRSSI(s):
    s=clear_between(s,"[DEBUG] "," ")
    time=find_between(s,"",",")
    rssi1=int(find_between(s,"0RSSI=",","))/2
    rssi2=int(find_between(s,"1RSSI=",","))/2
    cinr1=int(find_between(s,"0CINR=",","))/2
    cinr2=int(find_between(s,"1CINR=",","))/2
    uptime=find_between(s,"Uptime=",",")
    l=[time,rssi1,rssi2,cinr1,cinr2,uptime]
    return l

def reformWind(s,t):
    if t=="raw":
        s=clear_between(s,"[DEBUG] "," ")
        time=find_between(s,"",",")
        wind=find_between(s,"windSpeed=",",")
        l=[time,wind]
        return l
    else:
        s=clear_between(s,"[DEBUG] "," ")
        time=find_between(s,"",",")
        wind=find_between(s,"Wind speed average ",",")
        gust=find_between(s,"momentary gust ",s.strip()[-1])
        l=[time,wind,gust]
        return l


def reformBMS(s):
    s=clear_between(s,"[DEBUG] "," ")
    time=find_between(s,""," ")
    soc=find_between(s,"c=",",")
    soh=find_between(s,"h=",",")
    packVolt=find_between(s,"packVolt=",",")
    l=[time,soc,soh,packVolt]
    cells=find_between(s,"cellsVolts=[","]")
    cells=cells.split(",")
    l.extend(cells)
    for i in range(0,6):
        cells[i]=int(cells[i])
    diff=max(cells)-min(cells)
    l.append(diff)
    temp=find_between(s,"temperature=",",")
    l.append(temp)
    return l

def reformeDSTemp(s):
    date_time=find_between(s,"] "," [")
    date=date_time[0:10]
    time=date_time[10:-1]
    sensor=find_between(s,"p [","]")
    temp=clear_between(s,"[INFO"," ] : ")
    l=[date,time,sensor,temp]
    return l
 

if __name__ == "__main__":

    print('\n'.join(['Welcome!!!',
          '-------']))
    
##--------START ACTION: Browsing logs folder and setting output folders---------##
    
    root = tkinter.Tk()
    root.withdraw() #use to hide tkinter window
    currdir = os.getcwd()
    logsdir = filedialog.askdirectory(parent=root, initialdir=currdir, title='Please select a directory')
    if len(logsdir) > 0:
        print ("You chose " + logsdir)
    
    folderName=logsdir
    logsList=os.listdir(folderName)
    outdir = os.path.join(logsdir,"LogAnalyser Output")
    csvdir=os.path.join(outdir,"CSV Data Files")
    if not os.path.exists(csvdir):
        os.makedirs(outdir)
    if not os.path.exists(csvdir):
        os.makedirs(csvdir)
    
    
    #Files creation:
    #Start command csv file
    csvStep = open(os.path.join(csvdir,"Step.csv"),'w')
    step_wr = csv.writer(csvStep, dialect='excel',lineterminator='\n')
    csvStep.write("Time, Step\n")
        
    #RSSI/CINR csv file
    csvRSSI = open(os.path.join(csvdir,"Mobilicom.csv"),'w')
    rssi_wr = csv.writer(csvRSSI, dialect='excel',lineterminator='\n')
    csvRSSI.write("Time, RSSI 1, RSSI 2, CINR 1, CINR 2, Up Time, Step\n")

    #AvgWind csv file
    csvAvgWind = open(os.path.join(csvdir,"Average Wind.csv"),'w')
    wr_avgwind = csv.writer(csvAvgWind, dialect='excel',lineterminator='\n')
    csvAvgWind.write("Time, Wind, Gust, Step\n")

    #rawWind csv file
    csvRawWind = open(os.path.join(csvdir,"Raw Wind.csv"),'w')
    wr_rawwind = csv.writer(csvRawWind, dialect='excel',lineterminator='\n')
    csvRawWind.write("Time, Wind, Gust, Step\n")

    #BMS csv file
    csvBMS = open(os.path.join(csvdir,"BMS.csv"),'w')
    wr_bms = csv.writer(csvBMS, dialect='excel',lineterminator='\n')
    csvBMS.write("Time, SOC, SOH, Pack Volt, s1, s2, s3, s4 ,s5 ,s6, Difference, Temp., Step\n")

    #DStemp csv file
    csvDStemp = open(os.path.join(csvdir,"DS Temp.csv"),'w')
    wr_ds = csv.writer(csvDStemp, dialect='excel',lineterminator='\n')
    csvDStemp.write("Date, Time, Sensor, Temp\n")
    
    print("------\n"+
        "In progress !\n"+
        str(len(logsList))+
        " log files was imported\n"+
        "it going to take around "+
        str(round(5/16*len(logsList),2))+
        " seconds")
    
    csvcounter=0
    datacounter=0
    startAll=time.time()

##--------START ACTION: Getting data from log file and putting it in csvs---------##
    
    for file in logsList:
        path=os.path.join(folderName, file)
        #deals with each log file
        if os.path.isdir(path)==False:
            f = open(path,'r')
            contant = f.read().split("\n")
            step=[]
            flag=False
            for line in contant:
                #deals each line in log file
                rssicsvline=[]
                if "2: MobilicomGetSystemStatusResponse" in line:
                    rssicsvline=reformRSSI(line)
                    if flag:    
                        rssicsvline.append('>>'.join(step))
                        flag=False               
                    rssi_wr.writerow(rssicsvline)
                    csvcounter+=1
                elif "Wind speed average" in line:
                    csvline=reformWind(line,"avg")  
                    wr_avgwind.writerow(csvline)
                    csvcounter+=1
                elif "Message arrived Meteorology [mastId" in line:
                    csvline=reformWind(line,"raw")  
                    wr_rawwind.writerow(csvline)
                    csvcounter+=1
                elif "ARM message arrived ArmMessage [bmsMessage=BMSPeriodicMessage" in line:
                    csvline=reformBMS(line)
                    wr_bms.writerow(csvline)
                    csvcounter+=1
                elif "Start command" in line:
                    if not flag:
                        step=[]
                    step.append(reformStep(line)[1])
                    flag=True
                    step_wr.writerow(reformStep(line))
                    csvcounter+=1
                elif "updateAIlandingTempVariables()" in line:
                    csvline=reformeDSTemp(line)
                    wr_ds.writerow(csvline)
                    csvcounter+=1
                
                    
            datacounter+=len(contant)
            f.close()

    csvStep.close()
    csvRSSI.close()    
    csvAvgWind.close()
    csvRawWind.close()
    csvBMS.close()
    csvDStemp.close()

    dur=time.time()-startAll
    print("------\n"+
          "Data collection is DONE!\n"+
          str(len(logsList))+
          " files was imported\n"+
          "witch contains "+
          str(datacounter)+
          " lines, \n"+
          str(csvcounter)+
          " lines has parsed\n"+
          "It took " + str(round(dur,2)) +
          " sec to complete the task\n"+
          "------\n"
          "Marging CSV files into on Excel Workbook\n"
          "Please Wait!\n"
          "------")
    
##--------START ACTION: Margging CSV Files into one Excel file---------##
    
    writer = pd.ExcelWriter(os.path.join(outdir, 'Graphs.xlsx'))
    for filename in os.listdir(csvdir):
        if filename.split('.')[-1]=='csv':
            csvfile=os.path.join(csvdir, filename)
            df = pd.read_csv(csvfile)
            sheet_filename=filename.split("\\")[-1]
            sheet_filename=sheet_filename.split(".")[0]
            print(sheet_filename)
            df.to_excel(writer, sheet_name=sheet_filename)
                     
    writer.save()

    print("Graphs.xlsx has been created!\n"+
          "Restyling the data sheets\n"+
          "Please Wait!\n"+
          "------")


    wb = load_workbook(outdir+'/Graphs.xlsx')
    wsBMS=wb["BMS"]
    wsMob=wb["Mobilicom"]

    green_fill = PatternFill(start_color='8BC34A', end_color='8BC34A', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    red_fill = PatternFill(start_color='F44336', end_color='F44336', fill_type='solid')
    white_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
    white_font = Font(bold=True, color='ffffff')
    
    bms_rules = [CellIsRule(operator='between', formula=['0','100'], stopIfTrue=True, fill=green_fill),
                 CellIsRule(operator='between', formula=['100','200'], stopIfTrue=True, fill=orange_fill),
                 CellIsRule(operator='greaterThan', formula=['100'], stopIfTrue=True, fill=red_fill, font=white_font)]

    for rule in bms_rules:
        wsBMS.conditional_formatting.add('L:L',rule)

    rssi_rules = [CellIsRule(operator='between', formula=['-80','-90'], stopIfTrue=True, fill=orange_fill),
                  CellIsRule(operator='lessThan', formula=['-90'], stopIfTrue=True, fill=red_fill, font=white_font),
                  CellIsRule(operator='equal', formula=[''], stopIfTrue=True, fill=white_fill)]

    for rule in rssi_rules:
        wsMob.conditional_formatting.add('C:D',rule)

    cinr_rules = [CellIsRule(operator='between', formula=['6.1','7'], stopIfTrue=True, fill=orange_fill),
                  CellIsRule(operator='lessThan', formula=['6.1','-2'], stopIfTrue=True, fill=red_fill, font=white_font),
                  CellIsRule(operator='equal', formula=[''], stopIfTrue=True, fill=white_fill)]

    for rule in cinr_rules:
        wsMob.conditional_formatting.add('E:E',rule)
        wsMob.conditional_formatting.add('F:F',rule)

    
        
    wb.save(outdir+'/Graphs.xlsx')

    dur=time.time()-startAll
    print( "ALL DONE!!!\n"+
          "File name is:  Graphs.xlsx\n"+
          "All the process took " + str(round(dur,2))+
          "sec\n"
          "------\n")
   
      
    time.sleep(2)
##    print("Starting to plot the data\n"
##          "its going to take few seconds\n"
##          "------\n")

    

    



    
