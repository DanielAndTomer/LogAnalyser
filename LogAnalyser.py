import re
import csv
import os
import time
##import numpy as np
#import matplotlib.pyplot as plt
#from matplotlib import rc
import glob, xlwt
##import pandas as pd
import xlsxwriter
import tkinter
from tkinter import filedialog
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
##import plotly.plotly as py
##import plotly.graph_objs as go
##import plotly.figure_factory as FF


#rc('mathtext', default='regular')


def find_between( s, first, last ):
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""

def clear_between( s, first, last ):
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        toclear = s[start:end]
        toclear=first+toclear+last
        s=s.replace(toclear,"")
        return s
    except ValueError:
        print ("[ERROR]: Sub string didn't found")


def reformRSSI(s):
    s=clear_between(s,"[DEBUG] "," ")
    time=find_between(s,"",",")
    rssi1=int(find_between(s,"0RSSI=",","))/2
    rssi2=int(find_between(s,"1RSSI=",","))/2
    cinr1=int(find_between(s,"0CINR=",","))/2
    cinr2=int(find_between(s,"1CINR=",","))/2
    uptime=find_between(s,"Uptime=",",")
    l=[time,rssi1,rssi2,cinr1,cinr2,uptime]
    return l

def reformWind(s,t):
    if t=="raw":
        s=clear_between(s,"[DEBUG] "," ")
        time=find_between(s,"",",")
        wind=find_between(s,"windSpeed=",",")
        l=[time,wind]
        return l
    else:
        s=clear_between(s,"[DEBUG] "," ")
        time=find_between(s,"",",")
        wind=find_between(s,"Wind speed average ",",")
        gust=find_between(s,"momentary gust ",s.strip()[-1])
        l=[time,wind,gust]
        return l


def reformBMS(s):
    s=clear_between(s,"[DEBUG] "," ")
    time=find_between(s,""," ")
    soc=find_between(s,"c=",",")
    soh=find_between(s,"h=",",")
    packVolt=find_between(s,"packVolt=",",")
    l=[time,soc,soh,packVolt]
    cells=find_between(s,"cellsVolts=[","]")
    cells=cells.split(",")
    l.extend(cells)
    for i in range(0,6):
        cells[i]=int(cells[i])
    diff=max(cells)-min(cells)
    l.append(diff)
    temp=find_between(s,"temperature=",",")
    l.append(temp)
    return l

    

if __name__ == "__main__":

    print("Welcome!!!\n"
          "-------")

    root = tkinter.Tk()
    root.withdraw() #use to hide tkinter window
    currdir = os.getcwd()
    logsdir = filedialog.askdirectory(parent=root, initialdir=currdir, title='Please select a directory')
    if len(logsdir) > 0:
        print ("You chose " + logsdir)
    
    folderName=logsdir
    logsList=os.listdir(folderName)
    outdir = logsdir+"/LogAnalyser Output"
    if not os.path.exists(outdir):
        os.makedirs(outdir)
    
    #Files creation:
    #RSSI/CINR csv file
    csvRSSI = open(outdir+'/Mobilicom.csv','w')
    rssi_wr = csv.writer(csvRSSI, dialect='excel',lineterminator='\n')
    csvRSSI.write("Time, RSSI 1, RSSI 2, CINR 1, CINR 2, Up Time\n")

    #AvgWind csv file
    csvAvgWind = open(outdir+'/Average Wind.csv','w')
    wr_avgwind = csv.writer(csvAvgWind, dialect='excel',lineterminator='\n')
    csvAvgWind.write("Time, Wind, Gust\n")

    #rawWind csv file
    csvRawWind = open(outdir+'/Raw Wind.csv','w')
    wr_rawwind = csv.writer(csvRawWind, dialect='excel',lineterminator='\n')
    csvRawWind.write("Time, Wind, Gust\n")

    #BMS csv file
    csvBMS = open(outdir+'/BMS.csv','w')
    wr_bms = csv.writer(csvBMS, dialect='excel',lineterminator='\n')
    csvBMS.write("Time, SOC, SOH, Pack Volt, s1, s2, s3, s4 ,s5 ,s6, Difference, Temp.\n")
    
    print("------\n"+
        "In progress !\n"+
        str(len(logsList))+
        " log files was imported\n"+
        "it going to take around "+
        str(round(5/16*len(logsList),2))+
        " seconds")
    
    csvcounter=0
    datacounter=0
    startAll=time.time()
    
    for file in logsList:
        
        #deals with each log file
        if not file == folderName:
            path=folderName+"/"+file
            f = open(path,'r')
            contant = f.read().split("\n")
        
            for line in contant:
                #deals each line in log file
                if "2: MobilicomGetSystemStatusResponse" in line:
                    csvline=reformRSSI(line)
                    rssi_wr.writerow(csvline)
                    csvcounter+=1
                elif "Wind speed average" in line:
                    csvline=reformWind(line,"avg")
                    wr_avgwind.writerow(csvline)
                    csvcounter+=1
                elif "Message arrived Meteorology [mastId" in line:
                    csvline=reformWind(line,"raw")
                    wr_rawwind.writerow(csvline)
                    csvcounter+=1
                elif "ARM message arrived ArmMessage [bmsMessage=BMSPeriodicMessage" in line:
                    csvline=reformBMS(line)
                    wr_bms.writerow(csvline)
                    csvcounter+=1
                    
            datacounter+=len(contant)
            f.close()

    csvRSSI.close()    
    csvAvgWind.close()
    csvRawWind.close()
    csvBMS.close()

    dur=time.time()-startAll
    print("------\n"+
          "Data collection is DONE!\n"+
          str(len(logsList))+
          " files was imported\n"+
          "witch contains "+
          str(datacounter)+
          " lines, \n"+
          str(csvcounter)+
          " lines has parsed\n"+
          "It took " + str(round(dur,2)) +
          " sec to complete the task\n"+
          "------\n"
          "Marging CSV files into on Excel Workbook\n"
          "Please Wait!\n"
          "------")


##    df = pd.read_csv(outdir+'/Mobilicom.csv')
##
##    sample_data_table = FF.create_table(df.head())
##    py.iplot(sample_data_table, filename='Mobilicom_table')

    
##    trace = go.Scatter(x = df['Time'], y = df['RSSI 1'],
##                      name='TOMER')
##    layout = go.Layout(title='RSSI',
##                       plot_bgcolor='rgb(230, 230,230)', 
##                       showlegend=True)
##    fig = go.Figure(data=[trace], layout=layout)
##
##    py.iplot(fig, filename='tomer')
    
##    writer = pd.ExcelWriter(outdir+'/Graphs.xlsx')
##    for filename in glob.glob(outdir+'/*.csv'):
##        df = pd.read_csv(filename)
##        sheet_filename=filename.split("\\")[-1]
##        sheet_filename=sheet_filename.split(".")[0]
##        print(sheet_filename)
##        df.to_excel(writer, sheet_name=sheet_filename)
##        
##        os.remove(filename)
##    writer.save()
##
##    print("Graphs.xlsx has been created!\n"+
##          "Restyling the data sheets\n"+
##          "Please Wait!\n"+
##          "------")
##
##
##    wb = load_workbook(outdir+'/Graphs.xlsx')
##    wsBMS=wb["BMS"]
##    wsMob=wb["Mobilicom"]
##
##    green_fill = PatternFill(start_color='8BC34A', end_color='8BC34A', fill_type='solid')
##    orange_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
##    red_fill = PatternFill(start_color='F44336', end_color='F44336', fill_type='solid')
##    white_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
##    white_font = Font(bold=True, color='ffffff')
##    
##    bms_rules = [CellIsRule(operator='between', formula=['0','100'], stopIfTrue=True, fill=green_fill),
##                 CellIsRule(operator='between', formula=['100','200'], stopIfTrue=True, fill=orange_fill),
##                 CellIsRule(operator='greaterThan', formula=['100'], stopIfTrue=True, fill=red_fill, font=white_font)]
##
##    for rule in bms_rules:
##        wsBMS.conditional_formatting.add('L:L',rule)
##
##    rssi_rules = [CellIsRule(operator='between', formula=['-80','-90'], stopIfTrue=True, fill=orange_fill),
##                  CellIsRule(operator='lessThan', formula=['-90'], stopIfTrue=True, fill=red_fill, font=white_font)]
##
##    for rule in rssi_rules:
##        wsMob.conditional_formatting.add('C:D'.format(wsMob.max_row),rule)
##
##    cinr_rules = [CellIsRule(operator='between', formula=['6.1','7'], stopIfTrue=True, fill=orange_fill),
##                  CellIsRule(operator='lessThan', formula=['6.1','-2'], stopIfTrue=True, fill=red_fill, font=white_font)]
##
##    for rule in cinr_rules:
##        wsMob.conditional_formatting.add('E:F',rule)
##
##    
##        
##    wb.save(outdir+'/Graphs.xlsx')
##
##    dur=time.time()-startAll
##    print( "ALL DONE!!!\n"+
##          "File name is:  Graphs.xlsx\n"+
##          "All the process took " + str(round(dur,2))+
##          "sec\n"
##          "------\n")
##   
##      
##    time.sleep(2)
####    print("Starting to plot the data\n"
####          "its going to take few seconds\n"
####          "------\n")

    

    



    
